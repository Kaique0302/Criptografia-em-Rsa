import os
import random
from math import gcd
import logging
import openpyxl

# Configuração do logger
logging.basicConfig(
    filename='rsa_log.txt',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def log_message(message):
    logging.info(message)

# Função para criar a pasta 'chaves' e 'mensagens'
def criar_pasta_chaves():
    os.makedirs('chaves', exist_ok=True)
    os.makedirs('mensagens', exist_ok=True)
    log_message("Pastas 'chaves' e 'mensagens' verificadas ou criadas.")

# Função para calcular modular exponencial
def calcular_modulo(base, expoente, modulo):
    return pow(base, expoente, modulo)

# Função para calcular o inverso modular
def calcular_inverso_modular(expoente, totiente):
    valor, novo_valor = 0, 1
    resto, novo_resto = totiente, expoente
    while novo_resto:
        quociente = resto // novo_resto
        valor, novo_valor = novo_valor, valor - quociente * novo_valor
        resto, novo_resto = novo_resto, resto - quociente * novo_resto
    return valor + totiente if valor < 0 else valor

# Função para verificar se um número é primo
def verificar_primo(numero):
    if numero < 2:
        return False
    for i in range(2, int(numero**0.5) + 1):
        if numero % i == 0:
            return False
    return True

# Função para gerar chaves públicas e privadas
def gerar_chaves(primo_p, primo_q, expoente_publico, fator):
    modulo_n = primo_p * primo_q
    totiente_n = (primo_p - 1) * (primo_q - 1)
    expoente_privado = calcular_inverso_modular(expoente_publico, totiente_n)
    log_message(f"Chaves geradas: (N={modulo_n}, E={expoente_publico}), (N={modulo_n}, D={expoente_privado})")
    return (modulo_n, expoente_publico), (modulo_n, expoente_privado, fator)

#Ler mensagem encriptada
def ler_mensagem_encriptada():
    arquivo = 'mensagens/mensagem_encriptada.txt'
    try:
        with open(arquivo, 'r') as f:
            mensagem_encriptada = f.read()
            print("Mensagem criptografada:", mensagem_encriptada)
    except FileNotFoundError:
        print("Erro: O arquivo não foi encontrado.")

# Função para criptografar a mensagem
def criptografar(mensagem, expoente_publico, modulo_n, fator):
    mensagem_encriptada = ' '.join([str((calcular_modulo(byte, expoente_publico, modulo_n) * fator) % modulo_n) for byte in mensagem.encode('utf-8')])
    log_message(f"Mensagem criptografada com fator {fator}: {mensagem_encriptada}")
    
    # Salvar a mensagem encriptada em um arquivo
    with open('mensagens/mensagem_encriptada.txt', 'w') as f:
        f.write(mensagem_encriptada)
    
    return mensagem_encriptada

# Função para descriptografar a mensagem
def descriptografar(mensagem_encriptada, expoente_privado, modulo_n, fator):
    caracteres = []
    fator_inverso = pow(fator, -1, modulo_n)
    for codigo in mensagem_encriptada.split():
        char_code = calcular_modulo((int(codigo) * fator_inverso) % modulo_n, expoente_privado, modulo_n)
        caracteres.append(char_code)
    
    # Convertendo os códigos de bytes para uma string UTF-8
    mensagem = bytes(caracteres).decode('utf-8', errors='replace')
    log_message(f"Mensagem descriptografada: {mensagem}")
    
    # Salvar a mensagem desencriptada em um arquivo
    with open('mensagens/mensagem_desencriptada.txt', 'w') as f:
        f.write(mensagem)
    
    return mensagem

# Função de autenticação
def autenticar_usuario():
    pins_corretos = ["1308", "0816"]  # Altere para os PINs seguros de ambas as pessoas
    tentativas_restantes = 3

    while tentativas_restantes > 0:
        pin = input("Digite o PIN: ")
        if pin in pins_corretos:
            print("Autenticação bem-sucedida!")
            return True
        else:
            tentativas_restantes -= 1
            print(f"PIN incorreto. Tentativas restantes: {tentativas_restantes}")
            log_message("Tentativa de senha incorreta")
    
    print("Todas as tentativas esgotadas. Acesso negado.")
    return False

# Função para obter sugestões de números coprimos
def obter_sugestoes_coprimos(totiente):
    sugestoes = [i for i in range(2, totiente) if gcd(i, totiente) == 1]
    return random.sample(sugestoes, min(10, len(sugestoes)))

# Função para criptografar senhas em um arquivo .xlsx
def criptografar_senhas_arquivo():
    try:
        # Carregar o arquivo
        caminho_entrada = 'uploads/encriptarsenhas.xlsx'
        wb = openpyxl.load_workbook(caminho_entrada)
        sheet = wb.active
        
        # Solicitar chave pública
        modulo_n = int(input("Digite o valor de 'n' da chave pública: "))
        expoente_publico = int(input("Digite o valor de 'e' da chave pública: "))
        fator = int(input("Digite o fator para a criptografia: "))

        # Iterar pelas linhas e criptografar a coluna de senhas
        for row in range(2, sheet.max_row + 1):  # Começa em 2 para ignorar o cabeçalho
            senha = sheet.cell(row=row, column=2).value  # Supondo que a coluna 2 é de senhas
            if senha:
                senha_encriptada = criptografar(senha, expoente_publico, modulo_n, fator)
                sheet.cell(row=row, column=2).value = senha_encriptada  # Substitui a senha pela encriptada

        # Salvar o novo arquivo
        caminho_saida = 'uploads/encriptadosenhas.xlsx'
        wb.save(caminho_saida)
        print(f"As senhas foram criptografadas e salvas em '{caminho_saida}'.")

    except Exception as e:
        print(f"Ocorreu um erro: {e}")

# Função para desencriptar senhas em um arquivo .xlsx
def desencriptar_senhas_arquivo():
    try:
        # Carregar o arquivo
        caminho_entrada = 'uploads/encriptadosenhas.xlsx'
        wb = openpyxl.load_workbook(caminho_entrada)
        sheet = wb.active
        
        # Solicitar chave privada
        modulo_n = int(input("Digite o valor de 'n' da chave privada: "))
        expoente_privado = int(input("Digite o valor de 'd' da chave privada: "))
        fator = int(input("Digite o fator usado para a criptografia: "))

        # Iterar pelas linhas e desencriptar a coluna de senhas
        for row in range(2, sheet.max_row + 1):  # Começa em 2 para ignorar o cabeçalho
            senha_encriptada = sheet.cell(row=row, column=2).value  # Supondo que a coluna 2 é de senhas
            if senha_encriptada:
                senha = descriptografar(senha_encriptada, expoente_privado, modulo_n, fator)
                sheet.cell(row=row, column=2).value = senha  # Substitui a senha pela desencriptada

        # Salvar o novo arquivo
        caminho_saida = 'uploads/senhas_desencriptadas.xlsx'
        wb.save(caminho_saida)
        print(f"As senhas foram desencriptadas e salvas em '{caminho_saida}'.")

    except Exception as e:
        print(f"Ocorreu um erro: {e}")

# Função principal
def main():
    criar_pasta_chaves()

    # Autenticação do usuário
    if not autenticar_usuario():
        return

    opcao = None
    while opcao != "0":
        print("\nEscolha uma opção:")
        print("1 - Gerar chave pública e privada")
        print("2 - Criptografar mensagem")
        print("3 - Descriptografar mensagem")
        print("4 - Criptografar senhas de arquivo")
        print("5 - Desencriptar senhas de arquivo")
        print("0 - Sair")
        opcao = input("Opção: ")

        if opcao == "1":
            try:
                primo_p = int(input("Digite o número primo 'p': "))
                primo_q = int(input("Digite o número primo 'q': "))

                modulo_n = primo_p * primo_q
                totiente_n = (primo_p - 1) * (primo_q - 1)
                
                # Verificar se 'p' e 'q' são primos
                if not (verificar_primo(primo_p) and verificar_primo(primo_q)):
                    log_message("Um ou ambos os números fornecidos não são primos.")
                    print("Por favor, insira números primos válidos.")
                    continue

                fator = int(input("Digite um fator de segurança: "))

                # Sugestões de coprimos
                sugestoes = obter_sugestoes_coprimos(totiente_n)
                print(f"Sugestões de números coprimos: {sugestoes}")

                expoente_publico = int(input("Digite o expoente público 'e': "))

                # Verificação se 'e' é coprimo com φ(n)
                if gcd(expoente_publico, totiente_n) != 1:
                    print("O valor não é coprimo e pode retornar problemas na criptografia.")
                    continue
                
                # Gerar chaves
                chaves_publicas, chaves_privadas = gerar_chaves(primo_p, primo_q, expoente_publico, fator)
                
                # Salvar chaves em arquivos
                with open(f'chaves/chave_publica.txt', 'w') as f:
                    f.write(f"N: {chaves_publicas[0]}, E: {chaves_publicas[1]}")
                with open(f'chaves/chave_privada.txt', 'w') as f:
                    f.write(f"N: {chaves_privadas[0]}, D: {chaves_privadas[1]}, fator: {chaves_privadas[2]}")
                
                print("Chaves geradas e salvas com sucesso.")
                log_message("Chaves salvas com sucesso.")

            except Exception as e:
                print(f"Ocorreu um erro: {e}")

        elif opcao == "2":
            mensagem = input("Digite a mensagem a ser criptografada: ")
            expoente_publico = int(input("Digite o valor de 'e' da chave pública: "))
            modulo_n = int(input("Digite o valor de 'n' da chave pública: "))
            fator = int(input("Digite o fator para a criptografia: "))
            criptografar(mensagem, expoente_publico, modulo_n, fator)
            ler_mensagem_encriptada()

        elif opcao == "3":
            # Ler mensagem encriptada do arquivo
            try:
                with open('mensagens/mensagem_encriptada.txt', 'r') as f:
                    mensagem_encriptada = f.read().strip()
                expoente_privado = int(input("Digite o valor de 'd' da chave privada: "))
                modulo_n = int(input("Digite o valor de 'n' da chave privada: "))
                fator = int(input("Digite o fator usado para a criptografia: "))
                mensagem = descriptografar(mensagem_encriptada, expoente_privado, modulo_n, fator)
                print(f"Mensagem desencriptada: {mensagem}")
            except FileNotFoundError:
                print("Arquivo de mensagem encriptada não encontrado.")

        elif opcao == "4":
            criptografar_senhas_arquivo()

        elif opcao == "5":
            desencriptar_senhas_arquivo()

        elif opcao == "0":
            print("Saindo...")
        else:
            print("Opção inválida!")

if __name__ == "__main__":
    main()
